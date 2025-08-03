import streamlit as st
import google.generativeai as genai
import pdfplumber
import PyPDF2
import os
import json
from datetime import datetime
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from collections import Counter
import re
from sentence_transformers import SentenceTransformer
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
import tiktoken
from Sastrawi.StopWordRemover.StopWordRemoverFactory import StopWordRemoverFactory
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
import base64
import pickle
import hashlib
from wordcloud import WordCloud
import random
from PIL import Image
import openpyxl
from openpyxl import load_workbook
import csv
import time

# Load icon
try:
    icon = Image.open("assets/logo_bps.png")
except:
    icon = None

# Konfigurasi halaman
st.set_page_config(
    page_title="Ruwai Jurai - Interaksi Warga BPS Lampung",
    page_icon=icon,
    layout="wide"
)

# Konfigurasi Gemini API 
try:
    genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
except:
    st.error("GEMINI_API_KEY tidak ditemukan dalam secrets. Pastikan API key sudah dikonfigurasi.")
    st.stop()

# PERBAIKAN: Menggunakan st.session_state untuk cache persisten
CACHE_VERSION = "v2.0"  # Increment ketika struktur cache berubah

# Path ke folder dokumen
DOCUMENTS_FOLDER_PATH = r"files"

# File untuk menyimpan log chat
CHAT_LOG_FILE = "chat_logs.json"

# Token limits
MAX_USER_MESSAGE_TOKENS = 1500
MAX_SESSION_TOKENS = 10000

# Indonesian stopwords - comprehensive list
INDONESIAN_STOPWORDS = {
    # Kata hubung
    'dan', 'atau', 'tetapi', 'namun', 'serta', 'kemudian', 'lalu', 'setelah', 'sebelum',
    'ketika', 'saat', 'sambil', 'selama', 'hingga', 'sampai', 'karena', 'sebab',
    'oleh', 'karena', 'akibat', 'supaya', 'agar', 'untuk', 'demi', 'guna',
    
    # Kata depan
    'di', 'ke', 'dari', 'pada', 'dalam', 'dengan', 'oleh', 'bagi', 'untuk',
    'tentang', 'mengenai', 'terhadap', 'atas', 'bawah', 'antara', 'antar',
    'selain', 'kecuali', 'hingga', 'sampai', 'sejak', 'semenjak',
    
    # Kata ganti
    'saya', 'aku', 'kamu', 'anda', 'dia', 'ia', 'mereka', 'kita', 'kami',
    'ini', 'itu', 'tersebut', 'berikut', 'yang', 'mana', 'siapa', 'apa',
    'dimana', 'kemana', 'darimana', 'bagaimana', 'mengapa', 'kenapa',
    'kapan', 'bilamana', 'berapa', 'seberapa',
    
    # Kata kerja bantu
    'adalah', 'ialah', 'merupakan', 'yakni', 'yaitu', 'akan', 'sedang',
    'telah', 'sudah', 'pernah', 'belum', 'masih', 'sempat', 'baru',
    'dapat', 'bisa', 'mampu', 'sanggup', 'mau', 'ingin', 'hendak',
    'harus', 'wajib', 'perlu', 'butuh', 'boleh', 'jangan', 'tidak',
    'tak', 'bukan', 'belum', 'tanpa', 'kecuali', 'selain',
    
    # Kata keterangan
    'sangat', 'amat', 'sekali', 'banget', 'terlalu', 'cukup', 'agak',
    'sedikit', 'banyak', 'seluruh', 'semua', 'selalu', 'sering',
    'jarang', 'kadang', 'pernah', 'tidak', 'juga', 'pula', 'lagi',
    'masih', 'sudah', 'belum', 'baru', 'lama', 'sekarang', 'kini',
    'nanti', 'besok', 'kemarin', 'tadi', 'dulu', 'dahulu', 'lampau',
    'mendatang', 'akan', 'bakal', 'segera', 'langsung', 'seketika',
    
    # Kata umum lainnya yang sering muncul
    'yg', 'dgn', 'utk', 'dg', 'ttg', 'tsb', 'krn', 'pd', 'tdk',
    'gan', 'min', 'bang', 'bro', 'sis', 'om', 'tante', 'kak', 'dek',
    'wkwk', 'wkwkwk', 'haha', 'hihi', 'hehe', 'lol', 'wow', 'mantap',
    'oke', 'ok', 'thanks', 'thank', 'you', 'makasih', 'terima', 'kasih',
    'ada', 'mana', 'jadi', 'begitu', 'seperti', 'ibarat', 'bagai',
    'seolah', 'seakan', 'seumpama', 'umpama', 'misalnya', 'contohnya',
}

def get_indonesian_stopwords():
    """Mendapatkan stopwords bahasa Indonesia menggunakan Sastrawi"""
    try:
        # Menggunakan Sastrawi untuk stopwords yang lebih lengkap
        factory = StopWordRemoverFactory()
        stopwords = factory.get_stop_words()
        
        # Tambahkan stopwords manual jika diperlukan
        additional_stopwords = INDONESIAN_STOPWORDS
        
        # Gabungkan dengan stopwords tambahan
        all_stopwords = set(stopwords) | additional_stopwords
        return all_stopwords
        
    except ImportError:
        # Jika Sastrawi tidak tersedia, gunakan stopwords manual
        return INDONESIAN_STOPWORDS

# Token Counter Class
class TokenCounter:
    """Class untuk menghitung token dalam bahasa Indonesia"""
    
    @staticmethod
    def count_tokens(text):
        """Menghitung token dengan perkiraan untuk bahasa Indonesia"""
        if not text:
            return 0
        # Perkiraan: 1 token ≈ 3-4 karakter untuk bahasa Indonesia
        return len(text) // 3
    
    @staticmethod
    def get_session_tokens(messages):
        """Menghitung total token dalam sesi chat"""
        total_tokens = 0
        for message in messages:
            total_tokens += TokenCounter.count_tokens(message["content"])
        return total_tokens
    
    @staticmethod
    def trim_session_to_limit(messages, max_tokens):
        """Memotong pesan lama jika melebihi batas token"""
        if not messages:
            return messages
            
        # Hitung token dari belakang
        total_tokens = 0
        trimmed_messages = []
        
        # Mulai dari pesan terbaru
        for message in reversed(messages):
            message_tokens = TokenCounter.count_tokens(message["content"])
            if total_tokens + message_tokens <= max_tokens:
                trimmed_messages.insert(0, message)
                total_tokens += message_tokens
            else:
                break
        
        return trimmed_messages

class PersistentDocumentProcessor:
    """PERBAIKAN: Document processor dengan persistent cache menggunakan session_state"""
    
    def __init__(self, documents_folder_path):
        self.documents_folder_path = documents_folder_path
        self.chunk_size = 1000
        self.max_context_length = 30000
        
        # PERBAIKAN: Inisialisasi dengan persistent cache
        self._init_persistent_cache()
        
        # Load atau process documents
        self.load_documents()
    
    def _init_persistent_cache(self):
        """Inisialisasi cache persisten dalam session_state"""
        cache_key = f"document_cache_{CACHE_VERSION}"
        
        if cache_key not in st.session_state:
            st.session_state[cache_key] = {
                'document_contents': {},
                'document_chunks': {},
                'document_hashes': {},
                'table_contexts': {},
                'last_check': None,
                'is_initialized': False
            }
        
        # Assign references untuk akses mudah
        self.cache = st.session_state[cache_key]
        self.document_contents = self.cache['document_contents']
        self.document_chunks = self.cache['document_chunks']
        self.document_hashes = self.cache['document_hashes']
        self.table_contexts = self.cache['table_contexts']
    
    def get_file_hash(self, file_path):
        """Menghitung hash dari satu file berdasarkan modified time dan size"""
        try:
            file_stat = os.stat(file_path)
            file_info = f"{os.path.basename(file_path)}_{file_stat.st_mtime}_{file_stat.st_size}"
            return hashlib.md5(file_info.encode('utf-8')).hexdigest()
        except:
            return None
    
    def get_all_document_files(self):
        """Mendapatkan semua file dokumen yang didukung"""
        if not os.path.exists(self.documents_folder_path):
            return {}
        
        supported_extensions = ['.pdf', '.xlsx', '.xls', '.csv']
        document_files = {}
        
        for filename in os.listdir(self.documents_folder_path):
            if any(filename.lower().endswith(ext) for ext in supported_extensions):
                file_path = os.path.join(self.documents_folder_path, filename)
                file_hash = self.get_file_hash(file_path)
                if file_hash:
                    document_files[filename] = {
                        'path': file_path,
                        'hash': file_hash
                    }
        
        return document_files
    
    def load_documents(self):
        """PERBAIKAN: Load dokumen dengan checking minimal untuk performa optimal"""
        
        # Jika sudah diinisialisasi dan tidak perlu check ulang
        current_time = time.time()
        if (self.cache['is_initialized'] and 
            self.cache['last_check'] and 
            current_time - self.cache['last_check'] < 300):  # Check setiap 5 menit
            return
        
        # Get current files
        current_files = self.get_all_document_files()
        
        if not current_files:
            if not self.cache['is_initialized']:
                st.warning("⚠️ Tidak ada file dokumen yang ditemukan dalam folder")
            return
        
        # PERBAIKAN: Hanya process file yang baru atau berubah
        files_to_process = []
        files_updated = []
        files_new = []
        
        for filename, file_info in current_files.items():
            cached_hash = self.document_hashes.get(filename)
            current_hash = file_info['hash']
            
            if cached_hash != current_hash:
                if cached_hash is None:
                    files_new.append((filename, file_info))
                else:
                    files_updated.append((filename, file_info))
                files_to_process.append((filename, file_info))
        
        # Remove deleted files from cache
        files_to_remove = []
        for filename in list(self.document_hashes.keys()):
            if filename not in current_files:
                files_to_remove.append(filename)
        
        for filename in files_to_remove:
            self._remove_file_from_cache(filename)
        
        # Display status hanya jika ada perubahan atau belum diinisialisasi
        if not self.cache['is_initialized']:
            if self.document_contents:
                st.success(f"✅ Cache ditemukan: {len(self.document_contents)} dokumen")
            
            if files_to_process:
                st.info(f"🔄 Memproses {len(files_to_process)} dokumen...")
                self._process_files(files_to_process)
        
        elif files_to_process:
            # Ada file baru atau berubah
            status_msg = []
            if files_new:
                status_msg.append(f"{len(files_new)} file baru")
            if files_updated:
                status_msg.append(f"{len(files_updated)} file diperbarui")
            
            st.info(f"🔄 Memproses {', '.join(status_msg)}...")
            self._process_files(files_to_process)
        
        # Update cache metadata
        self.cache['last_check'] = current_time
        self.cache['is_initialized'] = True
        
        # Summary
        if files_to_process or not self.cache['is_initialized']:
            total_docs = len(self.document_contents)
            total_chunks = sum(len(chunks) for chunks in self.document_chunks.values())
            st.success(f"✅ Siap: {total_docs} dokumen, {total_chunks} chunks")
    
    def _process_files(self, files_to_process):
        """Process multiple files dengan progress bar"""
        if not files_to_process:
            return
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, (filename, file_info) in enumerate(files_to_process):
            status_text.text(f"Memproses {filename}...")
            
            success = self.process_single_document(filename, file_info['path'])
            
            if success:
                # Update hash setelah berhasil diproses
                self.document_hashes[filename] = file_info['hash']
                
            progress_bar.progress((i + 1) / len(files_to_process))
        
        status_text.empty()
        progress_bar.empty()
    
    def _remove_file_from_cache(self, filename):
        """Menghapus file dari cache"""
        if filename in self.document_contents:
            del self.document_contents[filename]
        if filename in self.document_chunks:
            del self.document_chunks[filename]
        if filename in self.document_hashes:
            del self.document_hashes[filename]
    
    def process_single_document(self, filename, file_path):
        """Memproses satu dokumen"""
        try:
            # Reset table_contexts untuk setiap dokumen
            self.table_contexts = {}
            
            text_content = None
            file_ext = filename.lower()
            
            if file_ext.endswith('.pdf'):
                text_content = self.extract_text_pdfplumber(file_path)
                if not text_content:
                    text_content = self.extract_text_pypdf2(file_path)
                    
            elif file_ext.endswith(('.xlsx', '.xls')):
                text_content = self.extract_text_excel(file_path)
                
            elif file_ext.endswith('.csv'):
                text_content = self.extract_text_csv(file_path)
            
            if text_content:
                # Clean and process text
                text_content = self.clean_text(text_content)
                self.document_contents[filename] = text_content
                
                # Create chunks
                chunks = self.chunk_text(text_content, self.chunk_size)
                self.document_chunks[filename] = chunks
                
                return True
            
            return False
            
        except Exception as e:
            st.error(f"Error processing {filename}: {str(e)}")
            return False
    
    def refresh_cache(self):
        """PERBAIKAN: Refresh cache dengan konfirmasi"""
        if st.button("⚠️ Konfirmasi Refresh Cache", type="secondary"):
            st.info("🔄 Menghapus cache dan memuat ulang semua dokumen...")
            
            # Clear cache
            self.cache['document_contents'].clear()
            self.cache['document_chunks'].clear()
            self.cache['document_hashes'].clear()
            self.cache['table_contexts'].clear()
            self.cache['is_initialized'] = False
            self.cache['last_check'] = None
            
            # Reload all documents
            self.load_documents()
            st.success("✅ Cache berhasil di-refresh!")
            st.rerun()
    
    def get_cache_info(self):
        """Mendapatkan informasi cache"""
        info = {
            'cache_exists': self.cache['is_initialized'],
            'cache_size': 0,
            'cached_at': self.cache.get('last_check'),
            'total_documents': len(self.document_contents),
            'total_chunks': sum(len(chunks) for chunks in self.document_chunks.values()),
            'cached_files': list(self.document_hashes.keys())
        }
        
        # Estimate cache size
        try:
            total_size = 0
            for content in self.document_contents.values():
                total_size += len(content.encode('utf-8'))
            info['cache_size'] = total_size
        except:
            pass
        
        return info
    
    def force_refresh_single_file(self, filename):
        """Force refresh satu file tertentu"""
        current_files = self.get_all_document_files()
        
        if filename in current_files:
            file_info = current_files[filename]
            
            # Remove from cache first
            self._remove_file_from_cache(filename)
            
            # Process again
            success = self.process_single_document(filename, file_info['path'])
            
            if success:
                self.document_hashes[filename] = file_info['hash']
                st.success(f"✅ {filename} berhasil di-refresh")
                return True
            else:
                st.error(f"❌ Gagal refresh {filename}")
                return False
        else:
            st.error(f"❌ File {filename} tidak ditemukan")
            return False
    
    def extract_text_excel(self, excel_path):
        """Ekstraksi teks dari file Excel (multiple sheets)"""
        try:
            text_content = ""
            workbook = load_workbook(excel_path, read_only=True, data_only=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_content += f"\n=== Sheet: {sheet_name} ===\n"
                
                # Baca header dulu (baris pertama)
                headers = []
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if first_row:
                    headers = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(first_row)]
                    text_content += "Headers: " + " | ".join(headers) + "\n\n"
                
                # Baca data (mulai dari baris kedua)
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if any(cell is not None for cell in row):  # Skip baris kosong
                        row_data = []
                        for i, cell in enumerate(row):
                            if i < len(headers):  # Pastikan tidak melebihi jumlah header
                                if cell is not None:
                                    row_data.append(f"{headers[i]}: {str(cell)}")
                        
                        if row_data:  # Jika ada data
                            text_content += f"Baris {row_num}: " + " | ".join(row_data) + "\n"
                
                text_content += "\n"
            
            workbook.close()
            return text_content
            
        except Exception as e:
            st.warning(f"Gagal membaca Excel {excel_path}: {str(e)}")
            return None
    
    def extract_text_csv(self, csv_path):
        """Ekstraksi teks dari file CSV"""
        try:
            text_content = ""
            
            encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
            
            for encoding in encodings:
                try:
                    with open(csv_path, 'r', encoding=encoding, newline='') as file:
                        sample = file.read(1024)
                        file.seek(0)
                        sniffer = csv.Sniffer()
                        delimiter = sniffer.sniff(sample).delimiter
                        
                        csv_reader = csv.DictReader(file, delimiter=delimiter)
                        
                        text_content += f"\n=== CSV File: {os.path.basename(csv_path)} ===\n"
                        
                        # Tampilkan header
                        headers = csv_reader.fieldnames
                        if headers:
                            text_content += "Headers: " + " | ".join(headers) + "\n\n"
                        
                        # Baca data
                        for row_num, row in enumerate(csv_reader, start=1):
                            row_data = []
                            for header, value in row.items():
                                if value and str(value).strip():
                                    row_data.append(f"{header}: {value}")
                            
                            if row_data:
                                text_content += f"Baris {row_num}: " + " | ".join(row_data) + "\n"
                        
                        break
                        
                except UnicodeDecodeError:
                    continue
                except Exception as e:
                    st.warning(f"Error reading CSV with {encoding}: {str(e)}")
                    continue
            
            return text_content if text_content.strip() else None
            
        except Exception as e:
            st.warning(f"Gagal membaca CSV {csv_path}: {str(e)}")
            return None
    
    def extract_text_pdfplumber(self, pdf_path):
        """Ekstraksi teks menggunakan pdfplumber - bekerja untuk semua jenis PDF"""
        try:
            text_content = ""
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    page_text = page.extract_text()
                    
                    # Analisis jenis halaman
                    page_type, table_info = self._analyze_page_type(page_text)
                    
                    if page_text:
                        # Tampilkan jenis halaman hanya jika relevan
                        if page_type == "Halaman Biasa":
                            text_content += f"\n--- Halaman {page_num} ---\n"
                        else:
                            text_content += f"\n--- Halaman {page_num} ({page_type}) ---\n"
                            
                            # Tambahkan referensi hanya untuk halaman lanjutan tabel
                            if page_type == "Lanjutan Tabel" and table_info and table_info in self.table_contexts:
                                original_info = self.table_contexts[table_info]
                                text_content += f"[REFERENSI: Melanjutkan {original_info['title']} dari Halaman {original_info['page']}]\n"
                        
                        text_content += page_text + "\n"
                    
                    # Ekstraksi tabel
                    tables = page.extract_tables()
                    if tables:
                        for table_num, table in enumerate(tables, 1):
                            # Tentukan nama tabel berdasarkan konteks
                            if page_type == "Tabel Utama" and table_info:
                                table_title = f"{table_info} (Halaman {page_num})"
                                # Simpan konteks untuk referensi masa depan
                                self.table_contexts[table_info] = {
                                    'title': table_info,
                                    'page': page_num,
                                    'headers': table[0] if table else None
                                }
                            elif page_type == "Lanjutan Tabel" and table_info and table_info in self.table_contexts:
                                original_title = self.table_contexts[table_info]['title']
                                table_title = f"Lanjutan {original_title} (Halaman {page_num})"
                            else:
                                table_title = f"Tabel {table_num} pada Halaman {page_num}"
                            
                            text_content += f"\n--- {table_title} ---\n"
                            
                            # Tampilkan header referensi jika diperlukan
                            if (page_type == "Lanjutan Tabel" and table_info and 
                                table_info in self.table_contexts and 
                                self._is_continuation_without_headers(table)):
                                
                                original_headers = self.table_contexts[table_info]['headers']
                                if original_headers:
                                    text_content += "[HEADER REFERENSI dari tabel asli:]\n"
                                    clean_headers = [str(cell) if cell is not None else "" for cell in original_headers]
                                    text_content += " | ".join(clean_headers) + "\n"
                                    text_content += "-" * 50 + "\n"
                            
                            # Ekstraksi data tabel
                            for row_num, row in enumerate(table):
                                if row:
                                    clean_row = [str(cell) if cell is not None else "" for cell in row]
                                    if row_num == 0 and self._looks_like_header(row):
                                        text_content += "[HEADER] " + " | ".join(clean_row) + "\n"
                                    else:
                                        text_content += " | ".join(clean_row) + "\n"
                            text_content += "\n"
            
            return text_content
        except Exception as e:
            st.warning(f"PDFPlumber gagal untuk {pdf_path}: {str(e)}")
            return None
    
    def _analyze_page_type(self, page_text):
        """Menganalisis jenis halaman berdasarkan teks"""
        if not page_text:
            return "Unknown", None
        
        patterns = {
            'continued_id': [
                r'Lanjutan\s+Tabel[/\\]?Continued\s+Table\s+(\d+(?:\.\d+)*)',
                r'Lanjutan\s+Tabel\s+(\d+(?:\.\d+)*)',
                r'Continued\s+Table\s+(\d+(?:\.\d+)*)',
                r'\(Lanjutan\)\s*Tabel\s+(\d+(?:\.\d+)*)',
                r'Tabel\s+(\d+(?:\.\d+)*)\s*\(Lanjutan\)',
            ],
            'main_table_id': [
                r'Tabel\s+(\d+(?:\.\d+)*)',
                r'Table\s+(\d+(?:\.\d+)*)',
                r'Lampiran\s+Tabel\s+(\d+(?:\.\d+)*)',
                r'Appendix\s+Table\s+(\d+(?:\.\d+)*)',
            ],
        }
        
        # Cek halaman lanjutan terlebih dahulu
        for pattern in patterns['continued_id']:
            match = re.search(pattern, page_text, re.IGNORECASE)
            if match:
                table_number = match.group(1)
                return "Lanjutan Tabel", f"Tabel {table_number}"
        
        # Cek tabel utama
        for pattern in patterns['main_table_id']:
            match = re.search(pattern, page_text, re.IGNORECASE)
            if match:
                table_number = match.group(1)
                return "Tabel Utama", f"Tabel {table_number}"
        
        return "Halaman Biasa", None
    
    def _looks_like_header(self, row):
        """Menentukan apakah baris terlihat seperti header"""
        if not row:
            return False
        
        header_indicators = ['tahun', 'year', 'jumlah', 'number', 'modal', 'investment', 
                           'proyek', 'project', 'sektor', 'sector', 'industri', 'industry']
        
        row_text = ' '.join([str(cell).lower() for cell in row if cell])
        return any(indicator in row_text for indicator in header_indicators)
    
    def _is_continuation_without_headers(self, table):
        """Menentukan apakah tabel lanjutan tidak memiliki header"""
        if not table or len(table) < 2:
            return True
        
        first_row = table[0]
        if first_row and len([cell for cell in first_row if cell and str(cell).isdigit()]) > len(first_row) / 2:
            return True
        
        return not self._looks_like_header(first_row)
    
    def extract_text_pypdf2(self, pdf_path):
        """Ekstraksi teks menggunakan PyPDF2 dengan penanganan lanjutan tabel"""
        try:
            text_content = ""
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    page_text = page.extract_text()
                    if page_text:
                        page_type, table_info = self._analyze_page_type(page_text)
                        
                        text_content += f"\n--- Halaman {page_num} ({page_type}) ---\n"
                        
                        if page_type == "Lanjutan Tabel" and table_info:
                            if table_info in self.table_contexts:
                                original_info = self.table_contexts[table_info]
                                text_content += f"[REFERENSI: Melanjutkan {original_info['title']} dari Halaman {original_info['page']}]\n"
                        
                        text_content += page_text + "\n"
                        
                        if page_type == "Tabel Utama" and table_info:
                            self.table_contexts[table_info] = {
                                'title': table_info,
                                'page': page_num,
                                'headers': None
                            }
            
            return text_content
        except Exception as e:
            st.warning(f"PyPDF2 gagal untuk {pdf_path}: {str(e)}")
            return None
    
    def chunk_text(self, text, chunk_size=1000, overlap=200):
        """Membagi teks menjadi chunk-chunk dengan overlap"""
        chunks = []
        start = 0
        text_length = len(text)
        
        while start < text_length:
            end = start + chunk_size
            
            if end < text_length:
                last_sentence = text.rfind('.', start, end)
                if last_sentence != -1 and last_sentence > start + chunk_size * 0.5:
                    end = last_sentence + 1
                else:
                    last_space = text.rfind(' ', start, end)
                    if last_space != -1 and last_space > start + chunk_size * 0.5:
                        end = last_space
            
            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)
            
            start = end - overlap
            if start >= text_length:
                break
        
        return chunks
    
    def clean_text(self, text):
        """Membersihkan teks dari karakter yang tidak perlu"""
        # Hapus karakter kontrol
        text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]', '', text)
        
        # Normalisasi whitespace
        text = re.sub(r'\s+', ' ', text)
        
        # Hapus baris kosong berlebihan
        text = re.sub(r'\n\s*\n', '\n\n', text)
        
        return text.strip()
    
    def count_tokens(self, text):
        """Menghitung token untuk model Gemini"""
        return TokenCounter.count_tokens(text)
    
    def search_relevant_content(self, query, max_chunks=10):
        """Mencari konten yang relevan"""
        relevant_chunks = []
        query_lower = query.lower()
        query_words = set(query_lower.split())
        
        for filename, chunks in self.document_chunks.items():
            for i, chunk in enumerate(chunks):
                chunk_lower = chunk.lower()
                chunk_words = set(chunk_lower.split())
                
                word_matches = len(query_words.intersection(chunk_words))
                exact_matches = sum(1 for word in query_words if word in chunk_lower)
                phrase_match = 1 if query_lower in chunk_lower else 0
                
                score = word_matches * 2 + exact_matches + phrase_match * 5
                
                if score > 0:
                    relevant_chunks.append({
                        'filename': filename,
                        'chunk_index': i,
                        'content': chunk,
                        'score': score
                    })
        
        relevant_chunks.sort(key=lambda x: x['score'], reverse=True)
        selected_chunks = relevant_chunks[:max_chunks]
        
        combined_content = ""
        total_tokens = 0
        
        for chunk_info in selected_chunks:
            chunk_text = f"\n=== Dari {chunk_info['filename']} (Bagian {chunk_info['chunk_index']+1}) ===\n"
            chunk_text += chunk_info['content'] + "\n"
            
            chunk_tokens = self.count_tokens(chunk_text)
            if total_tokens + chunk_tokens > self.max_context_length:
                break
            
            combined_content += chunk_text
            total_tokens += chunk_tokens
        
        if not combined_content and self.document_contents:
            first_file = list(self.document_contents.keys())[0]
            first_content = self.document_contents[first_file]
            
            if self.count_tokens(first_content) > self.max_context_length:
                chars_limit = self.max_context_length * 4
                combined_content = f"\n=== Dari {first_file} ===\n"
                combined_content += first_content[:chars_limit] + "...\n"
            else:
                combined_content = f"\n=== Dari {first_file} ===\n"
                combined_content += first_content
        
        return combined_content
    
    def get_all_content(self):
        """Menggabungkan semua konten dokumen dengan batasan token"""
        all_content = ""
        total_tokens = 0
        
        for filename, content in self.document_contents.items():
            file_content = f"\n=== Dokumen: {filename} ===\n{content}\n"
            file_tokens = self.count_tokens(file_content)
            
            if total_tokens + file_tokens > self.max_context_length:
                remaining_tokens = self.max_context_length - total_tokens
                remaining_chars = remaining_tokens * 4
                
                if remaining_chars > 500:
                    truncated_content = f"\n=== Dokumen: {filename} (Dipotong) ===\n"
                    truncated_content += content[:remaining_chars] + "...\n"
                    all_content += truncated_content
                break
            
            all_content += file_content
            total_tokens += file_tokens
        
        return all_content

class ImprovedWordCloudGenerator:
    def __init__(self, stopwords=None):
        self.stopwords = stopwords or get_indonesian_stopwords()
    
    def preprocess_text(self, text):
        """Preprocessing text untuk word cloud"""
        # Convert to lowercase
        text = text.lower()
        
        # Remove punctuation and numbers
        text = re.sub(r'[^\w\s]', ' ', text)
        text = re.sub(r'\d+', '', text)
        
        # Split into words
        words = text.split()
        
        # Filter stopwords dan kata pendek
        filtered_words = [
            word for word in words 
            if len(word) > 2 and word not in self.stopwords
        ]
        
        return ' '.join(filtered_words)
    
    def generate_wordcloud(self, text, max_words=100, width=800, height=400):
        """Generate word cloud dari text"""
        if not text.strip():
            return None
        
        # Preprocess text
        processed_text = self.preprocess_text(text)
        
        if not processed_text.strip():
            return None
        
        # Generate word cloud
        wordcloud = WordCloud(
            width=width,
            height=height,
            background_color='white',
            max_words=max_words,
            colormap='viridis',
            font_path=None,
            relative_scaling=0.5,
            min_font_size=10,
            max_font_size=100,
            prefer_horizontal=0.7,
            collocations=False
        ).generate(processed_text)
        
        return wordcloud
    
    def plot_wordcloud(self, wordcloud, title="Word Cloud"):
        """Plot word cloud menggunakan matplotlib"""
        fig, ax = plt.subplots(figsize=(12, 8))
        ax.imshow(wordcloud, interpolation='bilinear')
        ax.axis('off')
        ax.set_title(title, fontsize=16, fontweight='bold', pad=20)
        
        plt.tight_layout()
        
        return fig
    
    def get_word_frequencies(self, text, top_n=20):
        """Mendapatkan frekuensi kata untuk analisis lebih lanjut"""
        processed_text = self.preprocess_text(text)
        words = processed_text.split()
        
        if not words:
            return {}
        
        word_freq = Counter(words)
        return dict(word_freq.most_common(top_n))

class ChatLogger:
    def __init__(self, log_file):
        self.log_file = log_file
    
    def log_conversation(self, user_message, bot_response, is_document_related, tokens_used=0):
        """Menyimpan percakapan ke file log"""
        log_entry = {
            "timestamp": datetime.now().isoformat(),
            "user_message": user_message,
            "bot_response": bot_response,
            "is_document_related": is_document_related,
            "message_length": len(user_message),
            "response_length": len(bot_response),
            "tokens_used": tokens_used
        }
        
        # Baca log yang ada
        logs = self.load_logs()
        logs.append(log_entry)
        
        # Simpan ke file
        try:
            with open(self.log_file, 'w', encoding='utf-8') as f:
                json.dump(logs, f, ensure_ascii=False, indent=2)
        except Exception as e:
            st.error(f"Error saving log: {str(e)}")
    
    def load_logs(self):
        """Memuat log dari file"""
        if os.path.exists(self.log_file):
            try:
                with open(self.log_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return []
        return []

class GeminiChatbot:
    def __init__(self, document_processor):
        self.document_processor = document_processor
        try:
            self.model = genai.GenerativeModel('gemini-2.0-flash-exp')
        except Exception as e:
            st.error(f"Error initializing Gemini model: {str(e)}")
            self.model = None
        
    def is_document_related(self, message):
        """Menentukan apakah pertanyaan terkait dengan dokumen dengan logika yang lebih ketat"""
        # Kata kunci yang SANGAT spesifik untuk dokumen/statistik
        strong_document_keywords = [
            'dokumen', 'laporan', 'pdf', 'file', 'publikasi', 'tabel', 'grafik', 'excel', 'csv',
            'berdasarkan dokumen', 'dalam laporan', 'dari file', 'dari publikasi',
            'data statistik', 'sensus', 'survei bps', 'statistik lampung',
            'jumlah penduduk', 'data demografi', 'tingkat kemiskinan',
            'pdrb', 'inflasi', 'indeks', 'persentase penduduk', 'sheet', 'worksheet'
        ]
        
        # Kata kunci yang menunjukkan pertanyaan umum (bukan dokumen)
        general_keywords = [
            'halo', 'hai', 'hello', 'selamat', 'terima kasih', 'thanks',
            'bagaimana kabar', 'apa kabar', 'siapa kamu', 'nama kamu',
            'cara kerja', 'fungsi', 'pengertian', 'definisi', 'jelaskan',
            'maksud', 'arti', 'contoh', 'tips', 'saran', 'rekomendasi',
            'tutorial', 'panduan', 'langkah-langkah', 'prosedur'
        ]
        
        message_lower = message.lower()
        
        # Cek kata kunci umum dulu
        general_score = sum(1 for keyword in general_keywords if keyword in message_lower)
        if general_score >= 1 and len(message.split()) <= 10:
            return False
        
        # Cek kata kunci dokumen yang kuat
        strong_document_score = sum(2 for keyword in strong_document_keywords if keyword in message_lower)
        
        # Cek apakah ada angka/tahun yang menunjukkan pencarian data spesifik
        year_pattern = r'\b(19|20)\d{2}\b'
        number_pattern = r'\b\d+\b'
        has_year = bool(re.search(year_pattern, message))
        has_numbers = len(re.findall(number_pattern, message)) > 0
        
        # Cek konten dokumen relevan hanya jika ada indikasi kuat
        if strong_document_score >= 2 or has_year or (has_numbers and len(message.split()) > 5):
            relevant_content = self.document_processor.search_relevant_content(message)
            content_score = 3 if len(relevant_content.strip()) > 200 else 0
        else:
            content_score = 0
        
        # Scoring yang lebih ketat
        total_score = strong_document_score + content_score + (1 if has_year else 0)
        
        return total_score >= 3
    
    def generate_response(self, user_message):
        """Menghasilkan respons dari Gemini dengan handling yang lebih baik"""
        
        if not self.model:
            return "Maaf, terjadi kesalahan dengan model AI. Silakan coba lagi nanti.", False, 0
        
        # Handle greeting terlebih dahulu
        if self.is_greeting(user_message):
            greeting_responses = [
                "Halo! Saya adalah Asisten Virtual BPS Provinsi Lampung. Saya siap membantu Anda dengan informasi statistik dan pertanyaan lainnya. Ada yang bisa saya bantu?",
                "Selamat datang! Saya asisten virtual BPS Lampung. Silakan tanyakan apa saja yang ingin Anda ketahui tentang data statistik atau pertanyaan umum lainnya.",
                "Hai! Senang bertemu dengan Anda. Saya di sini untuk membantu menjawab pertanyaan tentang statistik Lampung atau topik lainnya. Bagaimana saya bisa membantu?"
            ]
            return random.choice(greeting_responses), False, 0
        
        # Handle thanks
        if self.is_thanks(user_message):
            thanks_responses = [
                "Sama-sama! Senang bisa membantu. Jangan ragu untuk bertanya lagi jika ada yang ingin Anda ketahui.",
                "Terima kasih kembali! Saya selalu siap membantu. Ada pertanyaan lain?",
                "Dengan senang hati! Semoga informasi yang saya berikan bermanfaat. Silakan bertanya kapan saja."
            ]
            return random.choice(thanks_responses), False, 0
        
        is_document_related = self.is_document_related(user_message)
        
        if is_document_related:
            # Jika terkait dokumen, gunakan konten dokumen sebagai konteks
            document_content = self.document_processor.search_relevant_content(user_message)
            tokens_used = self.document_processor.count_tokens(document_content)
            
            prompt = f"""
            Anda adalah Asisten Virtual BPS (Badan Pusat Statistik) Provinsi Lampung yang ahli dalam menganalisis data dan dokumen statistik.

            KONTEKS DOKUMEN:
            {document_content}

            PERTANYAAN PENGGUNA: {user_message}

            INSTRUKSI PENTING:
            1. Jawab pertanyaan berdasarkan HANYA informasi yang tersedia dalam dokumen di atas
            2. Jika ada tabel atau data numerik, presentasikan dengan format yang jelas dan rapi
            3. Jika informasi tidak tersedia dalam dokumen, jelaskan dengan sopan bahwa data tersebut tidak ada dalam dokumen yang tersedia
            4. Sebutkan sumber dokumen yang spesifik jika ada informasi dari beberapa dokumen (PDF, Excel, atau CSV)
            5. Gunakan bahasa Indonesia yang profesional namun mudah dipahami
            6. Jika ada data statistik, berikan konteks dan interpretasi yang helpful
            7. Jangan membuat asumsi atau menambahkan informasi yang tidak ada dalam dokumen

            FORMAT JAWABAN:
            - Mulai dengan jawaban langsung
            - Sertakan data/angka yang relevan jika ada
            - Berikan penjelasan atau interpretasi jika diperlukan
            - Sebutkan sumber dokumen di akhir
            """
        else:
            tokens_used = 0
            prompt = f"""
            Anda adalah Asisten Virtual BPS Provinsi Lampung yang membantu menjawab berbagai pertanyaan.

            PERTANYAAN PENGGUNA: {user_message}

            INSTRUKSI:
            1. Jawab pertanyaan dengan ramah, informatif, dan komprehensif menggunakan pengetahuan umum Anda
            2. Jika pertanyaan bersifat sapaan, balas dengan sapaan yang hangat
            3. Jika pertanyaan tentang definisi, konsep, atau penjelasan umum, berikan penjelasan yang mendalam dan jelas
            4. Jika pertanyaan teknis non-statistik, berikan jawaban yang komprehensif dengan contoh jika diperlukan
            5. Jika pertanyaan tentang sains, teknologi, sejarah, budaya, atau topik lainnya, gunakan pengetahuan Anda untuk memberikan jawaban yang akurat dan detail
            6. Gunakan bahasa Indonesia yang natural dan mudah dipahami
            7. Jika pertanyaan membutuhkan perhitungan atau analisis, lakukan dengan teliti
            8. Untuk pertanyaan tentang BPS atau statistik secara umum, berikan penjelasan yang edukatif dan komprehensif
            9. Jika user bertanya tentang Anda, jelaskan bahwa Anda adalah asisten virtual BPS Lampung yang dapat membantu dengan berbagai topik
            10. Berikan respons yang membantu, positif, dan menunjukkan kemampuan AI yang sesungguhnya
            11. Jika ada aspek pertanyaan yang kompleks, pecah menjadi bagian-bagian yang mudah dipahami
            12. Sertakan contoh, analogi, atau ilustrasi jika membantu pemahaman

            CATATAN: 
            - Manfaatkan seluruh pengetahuan dan kemampuan analisis Anda
            - Berikan jawaban yang komprehensif namun tetap mudah dipahami
            - Jangan batasi diri hanya pada topik statistik atau dokumen
            - Tunjukkan bahwa Anda adalah AI yang cerdas dan membantu
            """
        
        try:
            response = self.model.generate_content(prompt)
            return response.text, is_document_related, tokens_used
        except Exception as e:
            if is_document_related:
                error_msg = f"Maaf, terjadi kesalahan saat mengakses dokumen: {str(e)}"
            else:
                error_msg = f"Maaf, terjadi kesalahan saat memproses pertanyaan Anda: {str(e)}"
            return error_msg, is_document_related, tokens_used
    
    def is_greeting(self, message):
        """Mendeteksi apakah pesan adalah sapaan"""
        greetings = [
            'halo', 'hai', 'hello', 'selamat pagi', 'selamat siang', 
            'selamat sore', 'selamat malam', 'hi', 'hey', 'apa kabar',
            'bagaimana kabar', 'selamat datang'
        ]
        
        message_lower = message.lower().strip()
        return any(greeting in message_lower for greeting in greetings)

    def is_thanks(self, message):
        """Mendeteksi apakah pesan adalah ucapan terima kasih"""
        thanks_words = [
            'terima kasih', 'makasih', 'thanks', 'thank you', 'thx'
        ]
        
        message_lower = message.lower().strip()
        return any(thanks in message_lower for thanks in thanks_words)

def show_initial_greeting():
    """Menampilkan greeting otomatis saat pertama kali membuka chatbot"""
    if not st.session_state.greeting_shown and not st.session_state.messages:
        greeting_message = """
        🤖 **Halo! Selamat datang di Chatbot BPS Provinsi Lampung!**
        
        Saya adalah Asisten Virtual BPS (Badan Pusat Statistik) Provinsi Lampung yang siap membantu Anda dengan berbagai pertanyaan terkait statistik, data demografi, ekonomi, dan informasi lainnya yang relevan dengan Lampung.
        Silakan ketik pertanyaan Anda di bawah ini. Saya akan membantu mencari informasi dari dokumen yang tersedia atau memberikan penjelasan komprehensif berdasarkan pengetahuan saya.
        """
        
        st.session_state.messages.append({"role": "assistant", "content": greeting_message})
        st.session_state.greeting_shown = True

def init_session_state():
    """PERBAIKAN: Inisialisasi session state dengan cache persisten"""
    if 'messages' not in st.session_state:
        st.session_state.messages = []
    
    # PERBAIKAN: Inisialisasi document processor hanya sekali
    if 'document_processor' not in st.session_state:
        st.session_state.document_processor = PersistentDocumentProcessor(DOCUMENTS_FOLDER_PATH)
    
    if 'chatbot' not in st.session_state:
        st.session_state.chatbot = GeminiChatbot(st.session_state.document_processor)
    
    if 'chat_logger' not in st.session_state:
        st.session_state.chat_logger = ChatLogger(CHAT_LOG_FILE)
    
    if 'admin_authenticated' not in st.session_state:
        st.session_state.admin_authenticated = False
    
    if 'wordcloud_generator' not in st.session_state:
        st.session_state.wordcloud_generator = ImprovedWordCloudGenerator()
    
    if 'input_disabled' not in st.session_state:
        st.session_state.input_disabled = False
    
    if 'greeting_shown' not in st.session_state:
        st.session_state.greeting_shown = False

def enhanced_admin_cache_management():
    """Enhanced cache management interface untuk admin"""
    st.header("💾 Enhanced Cache Management")
    
    cache_info = st.session_state.document_processor.get_cache_info()
    
    # Overview metrics
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Cache Status", "✅ Aktif" if cache_info['cache_exists'] else "❌ Tidak Ada")
    
    with col2:
        cache_size_mb = cache_info['cache_size'] / (1024 * 1024) if cache_info['cache_size'] > 0 else 0
        st.metric("Cache Size", f"{cache_size_mb:.2f} MB")
    
    with col3:
        st.metric("Total Dokumen", cache_info['total_documents'])
    
    with col4:
        st.metric("Total Chunks", cache_info['total_chunks'])
    
    if cache_info['cached_at']:
        last_check = datetime.fromtimestamp(cache_info['cached_at'])
        st.info(f"📅 Cache terakhir diperiksa: {last_check.strftime('%Y-%m-%d %H:%M:%S')}")
    
    # File management section
    st.subheader("📁 File Management")
    
    if cache_info['cached_files']:
        # Create tabs for different views
        tab1, tab2, tab3 = st.tabs(["📋 File List", "🔄 Manual Operations", "📊 Cache Statistics"])
        
        with tab1:
            # Detailed file list
            st.write("**File yang tersimpan dalam cache:**")
            
            # Get current files from folder
            current_files = st.session_state.document_processor.get_all_document_files()
            
            for i, filename in enumerate(cache_info['cached_files']):
                with st.expander(f"📄 {filename}"):
                    col1, col2, col3 = st.columns([2, 1, 1])
                    
                    with col1:
                        # File info
                        if filename in current_files:
                            file_path = current_files[filename]['path']
                            try:
                                file_stat = os.stat(file_path)
                                file_size = file_stat.st_size / 1024  # KB
                                modified_time = datetime.fromtimestamp(file_stat.st_mtime)
                                
                                st.write(f"📊 Ukuran: {file_size:.1f} KB")
                                st.write(f"📅 Dimodifikasi: {modified_time.strftime('%Y-%m-%d %H:%M')}")
                                
                                # Show chunk info
                                if filename in st.session_state.document_processor.document_chunks:
                                    chunk_count = len(st.session_state.document_processor.document_chunks[filename])
                                    st.write(f"📄 Chunks: {chunk_count}")
                                
                            except Exception as e:
                                st.write(f"❌ Error reading file info: {str(e)}")
                        else:
                            st.write("⚠️ File tidak ditemukan di folder")
                    
                    with col2:
                        # Refresh single file
                        if st.button(f"🔄 Refresh", key=f"refresh_{i}"):
                            success = st.session_state.document_processor.force_refresh_single_file(filename)
                            if success:
                                st.rerun()
                    
                    with col3:
                        # Remove from cache
                        if st.button(f"🗑️ Remove", key=f"remove_{i}"):
                            st.session_state.document_processor._remove_file_from_cache(filename)
                            st.success(f"✅ {filename} dihapus dari cache")
                            st.rerun()
        
        with tab2:
            # Manual operations
            st.write("**Operasi Manual:**")
            
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("🔄 Check for Updates", help="Periksa file yang berubah"):
                    st.session_state.document_processor.load_documents()
                    st.rerun()
                
                if st.button("📊 Analyze Folder", help="Analisis perbedaan antara folder dan cache"):
                    current_files = st.session_state.document_processor.get_all_document_files()
                    cached_files = set(cache_info['cached_files'])
                    
                    st.write("**Analisis Folder vs Cache:**")
                    
                    # Files in folder but not cached
                    uncached = set(current_files.keys()) - cached_files
                    if uncached:
                        st.write(f"📁➡️💾 File di folder tapi belum di-cache: {len(uncached)}")
                        for f in uncached:
                            st.write(f"  • {f}")
                    
                    # Files cached but not in folder
                    missing = cached_files - set(current_files.keys())
                    if missing:
                        st.write(f"💾➡️❌ File di cache tapi tidak ada di folder: {len(missing)}")
                        for f in missing:
                            st.write(f"  • {f}")
                    
                    # Files that might need update
                    needs_update = []
                    for filename, file_info in current_files.items():
                        if filename in cached_files:
                            cached_hash = st.session_state.document_processor.document_hashes.get(filename)
                            if cached_hash != file_info['hash']:
                                needs_update.append(filename)
                    
                    if needs_update:
                        st.write(f"🔄 File yang perlu di-update: {len(needs_update)}")
                        for f in needs_update:
                            st.write(f"  • {f}")
                    
                    if not uncached and not missing and not needs_update:
                        st.success("✅ Cache dan folder sudah sinkron sempurna!")
            
            with col2:
                st.write("**Refresh Cache:**")
                st.session_state.document_processor.refresh_cache()
        
        with tab3:
            # Cache statistics
            st.write("**Cache Statistics:**")
            
            if cache_info['total_documents'] > 0:
                # File type distribution
                file_types = {}
                for filename in cache_info['cached_files']:
                    ext = os.path.splitext(filename)[1].lower()
                    file_types[ext] = file_types.get(ext, 0) + 1
                
                if file_types:
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.write("**Distribusi Jenis File:**")
                        for ext, count in file_types.items():
                            st.write(f"  • {ext}: {count} file(s)")
                    
                    with col2:
                        # Memory usage per file (estimated)
                        st.write("**Estimasi Penggunaan Memori:**")
                        for filename in cache_info['cached_files'][:5]:  # Show top 5
                            if filename in st.session_state.document_processor.document_contents:
                                content_size = len(st.session_state.document_processor.document_contents[filename])
                                st.write(f"  • {filename}: {content_size/1024:.1f} KB")
    else:
        st.info("Tidak ada file dalam cache. Silakan tambahkan dokumen ke folder dan refresh.")

def user_interface_alternative():
    """Interface alternatif dengan input form yang berpindah ke bawah saat processing"""
    st.title("🤖 Chatbot BPS Provinsi Lampung")
    st.markdown("Tanyakan apapun tentang data di Provinsi Lampung atau pertanyaan umum lainnya!")
    
    # Tampilkan greeting otomatis
    show_initial_greeting()
    
    # Hitung token sesi saat ini
    session_tokens = TokenCounter.get_session_tokens(st.session_state.messages)
    
    # Display token usage di sidebar
    st.sidebar.markdown("---")
    st.sidebar.markdown("📊 **Penggunaan Token**")
    
    # Progress bar untuk session tokens
    session_progress = min(session_tokens / MAX_SESSION_TOKENS, 1.0)
    st.sidebar.progress(session_progress)
    st.sidebar.write(f"Sesi: {session_tokens:,} / {MAX_SESSION_TOKENS:,} token")
    
    # Warning jika mendekati batas
    if session_tokens > MAX_SESSION_TOKENS * 0.8:
        st.sidebar.warning("⚠️ Sesi mendekati batas token. Pesan lama akan dihapus otomatis.")
    elif session_tokens > MAX_SESSION_TOKENS * 0.6:
        st.sidebar.info("ℹ️ Sesi telah menggunakan 60% dari batas token.")
    
    # Tambahkan contoh pertanyaan
    with st.expander("💡 Contoh Pertanyaan"):
        st.markdown("""
        **Pertanyaan tentang Data Statistik:**
        - Berapa jumlah penduduk Lampung tahun 2023?
        - Bagaimana tingkat kemiskinan di Lampung?
        - Data PDRB Lampung terbaru?
        
        **Pertanyaan Umum:**
        - Apa itu BPS?
        - Bagaimana cara menghitung inflasi?
        - Jelaskan tentang sensus penduduk
        - Apa fungsi statistik dalam pembangunan?
        """)
    
    # Container untuk chat history
    chat_container = st.container()
    
    # Tampilkan chat history
    with chat_container:
        for message in st.session_state.messages:
            if message["role"] == "user":
                with st.chat_message("user"):
                    st.write(message["content"])
            else:
                with st.chat_message("assistant"):
                    st.write(message["content"])
    
    # Tampilkan loading jika sedang processing
    if st.session_state.input_disabled:
        with st.spinner("🤖 Asisten sedang memproses pertanyaan Anda..."):
            user_input = st.session_state.current_user_input
            
            try:
                response, is_document_related, tokens_used = st.session_state.chatbot.generate_response(user_input)
                
                # Tambahkan respons bot
                st.session_state.messages.append({"role": "assistant", "content": response})
                
                # Log percakapan
                st.session_state.chat_logger.log_conversation(user_input, response, is_document_related, tokens_used)
                
            except Exception as e:
                st.error(f"Terjadi kesalahan: {str(e)}")
            
            finally:
                # Reset processing state
                st.session_state.input_disabled = False
                if hasattr(st.session_state, 'current_user_input'):
                    delattr(st.session_state, 'current_user_input')
                st.rerun()
    
    # Tampilkan input form hanya jika TIDAK sedang processing
    if not st.session_state.input_disabled:
        input_form()
    
    # Tampilkan statistik penggunaan
    if st.session_state.messages:
        st.sidebar.markdown("---")
        st.sidebar.markdown("📊 **Statistik Sesi**")
        st.sidebar.write(f"Total Pesan: {len(st.session_state.messages)}")
        user_messages = [msg for msg in st.session_state.messages if msg["role"] == "user"]
        st.sidebar.write(f"Pertanyaan Anda: {len(user_messages)}")

def input_form():
    """Form input yang bisa di-reuse"""
    with st.form("chat_form", clear_on_submit=True):
        user_input = st.text_area(
            "Ketik pertanyaan Anda di sini...", 
            height=100, 
            placeholder="Contoh: Berapa jumlah penduduk Lampung tahun 2023?",
            key="user_input_field"
        )
        
        # Hitung token dari input
        input_tokens = TokenCounter.count_tokens(user_input) if user_input else 0
        
        # Tampilkan jumlah token input
        if user_input:
            if input_tokens > MAX_USER_MESSAGE_TOKENS:
                st.error(f"❌ Pesan terlalu panjang! ({input_tokens:,} token, maksimal {MAX_USER_MESSAGE_TOKENS:,} token)")
                st.info("💡 Tip: Persingkat pertanyaan Anda atau bagi menjadi beberapa pertanyaan terpisah.")
            else:
                color = "green" if input_tokens <= MAX_USER_MESSAGE_TOKENS * 0.8 else "orange"
                st.markdown(f"<p style='color: {color}; font-size: 12px;'>Token: {input_tokens:,} / {MAX_USER_MESSAGE_TOKENS:,}</p>", unsafe_allow_html=True)
        
        col1, col2 = st.columns([3, 1])
        with col1:
            submitted = st.form_submit_button(
                "Kirim", 
                use_container_width=True,
                disabled=input_tokens > MAX_USER_MESSAGE_TOKENS
            )
        with col2:
            clear_chat = st.form_submit_button(
                "Bersihkan Chat", 
                use_container_width=True
            )
        
        # Handle form submission
        if clear_chat:
            st.session_state.messages = []
            st.rerun()
        
        if submitted and user_input and input_tokens <= MAX_USER_MESSAGE_TOKENS:
            # Tambahkan pesan pengguna
            st.session_state.messages.append({"role": "user", "content": user_input})
            
            # Trim pesan jika melebihi batas sesi
            updated_session_tokens = TokenCounter.get_session_tokens(st.session_state.messages)
            if updated_session_tokens > MAX_SESSION_TOKENS:
                st.session_state.messages = TokenCounter.trim_session_to_limit(
                    st.session_state.messages, MAX_SESSION_TOKENS
                )
                st.info("ℹ️ Beberapa pesan lama telah dihapus untuk menghemat token.")
            
            # Set processing state
            st.session_state.input_disabled = True
            st.session_state.current_user_input = user_input
            st.rerun()

def admin_login():
    """Halaman login admin"""
    st.title("🔐 Admin Login")
    
    with st.form("admin_login"):
        username = st.text_input("Username:")
        password = st.text_input("Password:", type="password")
        submit = st.form_submit_button("Login")
        
        if submit:
            # Ganti dengan credentials yang sesuai
            if username == "admin" and password == "admin123":
                st.session_state.admin_authenticated = True
                st.success("Login berhasil!")
                st.rerun()
            else:
                st.error("Username atau password salah!")
    
    # Tambahan info untuk user
    st.info("💡 Silakan login untuk mengakses dashboard admin")

def admin_interface():
    """Interface untuk admin dengan analytics dan word cloud"""
    st.title("🔧 Admin Dashboard - BPS Chatbot Analytics")
    
    # Load chat logs
    logs = st.session_state.chat_logger.load_logs()
    
    if not logs:
        st.warning("Belum ada data chat untuk dianalisis.")
        return
    
    # Sidebar untuk filter
    with st.sidebar:
        st.header("Filter Data")
        
        # Filter berdasarkan tanggal
        dates = [datetime.fromisoformat(log['timestamp']).date() for log in logs]
        if dates:
            min_date = min(dates)
            max_date = max(dates)
            
            start_date = st.date_input("Dari tanggal:", min_date)
            end_date = st.date_input("Sampai tanggal:", max_date)
            
            # Filter logs berdasarkan tanggal
            filtered_logs = [
                log for log in logs 
                if start_date <= datetime.fromisoformat(log['timestamp']).date() <= end_date
            ]
        else:
            filtered_logs = logs
        
        # Filter berdasarkan jenis pertanyaan
        question_types = st.multiselect(
            "Jenis Pertanyaan:",
            ["Document Related", "General"],
            default=["Document Related", "General"]
        )
        
        if question_types:
            filtered_logs = [
                log for log in filtered_logs
                if (log['is_document_related'] and "Document Related" in question_types) or
                   (not log['is_document_related'] and "General" in question_types)
            ]
    
    # Tabs untuk berbagai analisis
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Overview", "☁️ Word Cloud", "📈 Analytics", "💬 Chat History", "💾 Cache Management"])
    
    with tab1:
        # Overview metrics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Conversations", len(filtered_logs))
        
        with col2:
            document_related = sum(1 for log in filtered_logs if log['is_document_related'])
            st.metric("Document Related", document_related)
        
        with col3:
            general_questions = len(filtered_logs) - document_related
            st.metric("General Questions", general_questions)
        
        with col4:
            avg_response_length = sum(log['response_length'] for log in filtered_logs) / len(filtered_logs) if filtered_logs else 0
            st.metric("Avg Response Length", f"{avg_response_length:.0f} chars")
        
        # Chart distribusi per hari
        if filtered_logs:
            df_daily = pd.DataFrame(filtered_logs)
            df_daily['date'] = pd.to_datetime(df_daily['timestamp']).dt.date
            daily_counts = df_daily.groupby(['date', 'is_document_related']).size().reset_index(name='count')
            
            fig = px.bar(
                daily_counts, 
                x='date', 
                y='count',
                color='is_document_related',
                title="Distribusi Pertanyaan Harian",
                labels={'is_document_related': 'Document Related', 'count': 'Jumlah Pertanyaan'}
            )
            st.plotly_chart(fig, use_container_width=True)
    
    with tab2:
        # Word Cloud Analysis
        st.header("☁️ Word Cloud Analysis")
        
        if filtered_logs:
            # Gabungkan semua user messages
            all_user_messages = " ".join([log['user_message'] for log in filtered_logs])
            
            # Pilihan untuk word cloud
            col1, col2 = st.columns(2)
            
            with col1:
                max_words = st.slider("Maksimal Kata", 50, 200, 100)
                
            with col2:
                wordcloud_type = st.selectbox(
                    "Jenis Word Cloud:",
                    ["Semua Pertanyaan", "Document Related", "General Questions"]
                )
            
            # Filter berdasarkan jenis
            if wordcloud_type == "Document Related":
                messages = [log['user_message'] for log in filtered_logs if log['is_document_related']]
            elif wordcloud_type == "General Questions":
                messages = [log['user_message'] for log in filtered_logs if not log['is_document_related']]
            else:
                messages = [log['user_message'] for log in filtered_logs]
            
            if messages:
                combined_text = " ".join(messages)
                
                # Generate word cloud
                wordcloud = st.session_state.wordcloud_generator.generate_wordcloud(
                    combined_text, 
                    max_words=max_words,
                    width=1200,
                    height=600
                )
                
                if wordcloud:
                    # Plot word cloud
                    fig = st.session_state.wordcloud_generator.plot_wordcloud(
                        wordcloud, 
                        title=f"Word Cloud - {wordcloud_type}"
                    )
                    st.pyplot(fig)
                    
                    # Tampilkan frekuensi kata
                    st.subheader("📊 Frekuensi Kata Teratas")
                    word_freq = st.session_state.wordcloud_generator.get_word_frequencies(combined_text, top_n=20)
                    
                    if word_freq:
                        # Buat chart frekuensi
                        freq_df = pd.DataFrame(list(word_freq.items()), columns=['Kata', 'Frekuensi'])
                        fig_freq = px.bar(
                            freq_df, 
                            x='Frekuensi', 
                            y='Kata',
                            orientation='h',
                            title="Top 20 Kata yang Paling Sering Digunakan",
                            height=600
                        )
                        fig_freq.update_layout(yaxis={'categoryorder': 'total ascending'})
                        st.plotly_chart(fig_freq, use_container_width=True)
                        
                        # Tabel frekuensi
                        st.dataframe(freq_df, use_container_width=True)
                else:
                    st.warning("Tidak dapat membuat word cloud. Mungkin tidak ada kata yang cukup bermakna.")
            else:
                st.warning("Tidak ada data untuk jenis word cloud yang dipilih.")
        else:
            st.warning("Tidak ada data chat untuk membuat word cloud.")
    
    with tab3:
        # Advanced Analytics
        st.header("📈 Advanced Analytics")
        
        if filtered_logs:
            # Analisis panjang pesan
            st.subheader("Analisis Panjang Pesan")
            
            df_analysis = pd.DataFrame(filtered_logs)
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Histogram panjang user message
                fig_hist = px.histogram(
                    df_analysis, 
                    x='message_length',
                    nbins=20,
                    title="Distribusi Panjang Pertanyaan User"
                )
                st.plotly_chart(fig_hist, use_container_width=True)
            
            with col2:
                # Histogram panjang response
                fig_hist_resp = px.histogram(
                    df_analysis, 
                    x='response_length',
                    nbins=20,
                    title="Distribusi Panjang Respons Bot"
                )
                st.plotly_chart(fig_hist_resp, use_container_width=True)
            
            # Korelasi panjang pertanyaan vs respons
            st.subheader("Korelasi Panjang Pertanyaan vs Respons")
            fig_scatter = px.scatter(
                df_analysis,
                x='message_length',
                y='response_length',
                color='is_document_related',
                title="Korelasi Panjang Pertanyaan vs Panjang Respons",
                labels={'message_length': 'Panjang Pertanyaan', 'response_length': 'Panjang Respons'}
            )
            st.plotly_chart(fig_scatter, use_container_width=True)
    
    with tab4:
        # Chat History
        st.header("💬 Chat History")
        
        if filtered_logs:
            # Pagination
            items_per_page = st.selectbox("Items per page:", [10, 25, 50, 100], index=1)
            total_pages = (len(filtered_logs) + items_per_page - 1) // items_per_page
            
            if total_pages > 1:
                page = st.number_input("Page:", 1, total_pages, 1)
                start_idx = (page - 1) * items_per_page
                end_idx = start_idx + items_per_page
                page_logs = filtered_logs[start_idx:end_idx]
            else:
                page_logs = filtered_logs
            
            # Tampilkan chat history
            for i, log in enumerate(page_logs, 1):
                with st.expander(f"Chat {start_idx + i if 'start_idx' in locals() else i} - {log['timestamp'][:19]}"):
                    st.write("**User:**", log['user_message'])
                    st.write("**Bot:**", log['bot_response'])
                    st.write("**Document Related:**", "Ya" if log['is_document_related'] else "Tidak")
                    st.write("**Tokens Used:**", log.get('tokens_used', 'N/A'))
        else:
            st.info("Tidak ada chat history untuk ditampilkan.")
    
    with tab5:
        # Enhanced cache management
        enhanced_admin_cache_management()

def main():
    """Fungsi utama aplikasi"""
    init_session_state()
    
    # Sidebar untuk navigasi
    col1, col2, col3 = st.sidebar.columns([1, 2, 1])
    with col2:
        try:
            st.image("assets/logo_bps.png", width=100)
        except:
            st.write("🏛️ BPS")
    st.sidebar.title("Ruwai Jurai")
    st.sidebar.markdown("Ruang Interaksi Warga dengan BPS Provinsi Lampung")
    
    page = st.sidebar.selectbox("Pilih Halaman:", ["👤 User Chat", "👨‍💼 Admin Dashboard"])
    
    # Routing halaman
    if page == "👤 User Chat":
        user_interface_alternative()
    else:
        if not st.session_state.admin_authenticated:
            admin_login()
        else:
            st.sidebar.markdown("---")
            st.sidebar.markdown("👋 Selamat datang, Admin!")
            if st.sidebar.button("Logout"):
                st.session_state.admin_authenticated = False
                st.success("Anda telah logout")
                st.rerun()
            
            admin_interface()

# Utility functions for better performance
def clear_cache_on_demand():
    """Clear cache hanya saat diminta"""
    cache_key = f"document_cache_{CACHE_VERSION}"
    if cache_key in st.session_state:
        del st.session_state[cache_key]
    st.success("Cache berhasil dihapus!")

def get_cache_status():
    """Get status cache saat ini"""
    cache_key = f"document_cache_{CACHE_VERSION}"
    if cache_key in st.session_state:
        cache = st.session_state[cache_key]
        return {
            'is_initialized': cache.get('is_initialized', False),
            'document_count': len(cache.get('document_contents', {})),
            'last_check': cache.get('last_check'),
            'total_chunks': sum(len(chunks) for chunks in cache.get('document_chunks', {}).values())
        }
    return {'is_initialized': False, 'document_count': 0, 'last_check': None, 'total_chunks': 0}

# Performance monitoring untuk debug
def debug_performance():
    """Debug function untuk monitoring performa"""
    if st.sidebar.button("🔍 Debug Performance"):
        status = get_cache_status()
        st.sidebar.json(status)

# Main execution
if __name__ == "__main__":
    try:
        main()
        
        # PERBAIKAN: Tambahkan debug info di sidebar untuk development
        if st.sidebar.button("🔍 Show Cache Status"):
            status = get_cache_status()
            st.sidebar.json(status)
            
    except Exception as e:
        st.error(f"Terjadi kesalahan aplikasi: {str(e)}")
        st.info("Silakan refresh halaman atau hubungi administrator.")
        
        # PERBAIKAN: Tambahkan tombol untuk clear cache jika terjadi error
        if st.button("🗑️ Clear Cache & Restart"):
            clear_cache_on_demand()
            st.rerun()